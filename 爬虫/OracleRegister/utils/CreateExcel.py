# 使用openpyxl追加写入数据到Excel
import openpyxl
import os
import pandas as pd

def create_excel_xlsx(path, sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)

def write_excel_xlsx_append(path, value, truncate_sheet=False):
    # 如果不存在就创建该excel
    if not os.path.exists(path):
        create_excel_xlsx(path, 'Sheet1')

    value = value.values #将dataframe转为array
    data = openpyxl.load_workbook(path)
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if(truncate_sheet): #truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        # print(sheet.title)  # 输出表名
        startrows = sheet.max_row  # 获得行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=startrows + i + 1, column=j + 1, value=str(value[i][j]))
    data.save(path)
    print("xlsx格式表格追加写入数据成功！")

def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

# if __name__ == '__main__':
#
#     book_name_xlsx = 'xlsx格式测试工作簿.xlsx'
#     sheet_name_xlsx = 'xlsx格式测试表'
#     value3 = [["111", "女", "66", "石家庄", "运维工程师"],
#               ["222", "男", "55", "南京", "饭店老板"],
#               ["333", "女", "27", "苏州", "保安"], ]
#     write_excel_xlsx_append(book_name_xlsx, sheet_name_xlsx, value3)
#     read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

def main(book_name_xlsx,sheet_name_xlsx,value3):

    write_excel_xlsx_append(book_name_xlsx, sheet_name_xlsx, value3)
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

def test(fileName,data):
    # load_workbook() 一个参数,加载xlsl文件的路径,续写.
    wb = openpyxl.load_workbook(fileName)
    # 对xlsx文件操作之前,要 active.
    wa = wb.active
    # append() 一个参数,列表,写入的一行.
    wa.append(data)
    # save() 一个参数,保存路径
    wb.save(fileName)

if __name__ == '__main__':
    test1 = {
        '标题': [],
        '完整标题': [],
        '正文': []
    }
    df = pd.DataFrame(test1)
    df.to_excel('data_list.xlsx', index=False)